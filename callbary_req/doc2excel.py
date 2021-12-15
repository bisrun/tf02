import os.path
import docx

from docx import Document
from docx.document import Document as _Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph
from docx.opc.constants import RELATIONSHIP_TYPE as RT

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

wb = openpyxl.Workbook()
#워크북을 생성하면 그 안에 워크시트 1개가 자동으로 생성
ws = wb.active
# 활성화 된 워크시트를 가리킴
#wb.create_sheet('TITLE',0)
cols = ["CHAP1", "CHAP2","TITLE","REQID","내용","제약사항","비고"]
col_contents = ["", "", "", "", "", "", ""]
cols_idx = {"CHAP1":0,"CHAP2":1,"TITLE":2,"REQID":3,"내용":4,"제약사항":5,"비고":6}
cols_width = {"CHAP1":20,"CHAP2":20,"TITLE":30,"REQID":10,"내용":90,"제약사항":30,"비고":50}
detail_req_col_idx = -1

directory_path = 'D:/Documents/++ST++/00_PROOM2021/02.배송고도화/01.요구사항'
docx_filename = '화물중계서비스용물류엔진_요구사항_20211215_v14.docx'
docx_filepath = os.path.join(directory_path, docx_filename)
excel_filepath = docx_filepath.replace("docx","xlsx")

xrow = 1
for i, name in enumerate(cols) :
    ws.cell(row=xrow,column=i+1).value = name
    ws.column_dimensions[get_column_letter(cols_idx[name]+1)].width = cols_width[name]
xrow = xrow + 1


def writeExcelRow():
    for i, col in enumerate(col_contents):
        ws.cell(row=xrow,column=i+1).value = col

def read_doc():
    global xrow
    global col_contents

    directory_path = 'D:/Documents/++ST++/00_PROOM2021/02.배송고도화/01.요구사항'
    docx_filename = '화물중계서비스용물류엔진_요구사항_20211215_v13.docx'
    docx_filepath = os.path.join(directory_path, docx_filename)
    excel_filepath = docx_filepath.replace("docx","xlsx")

    chapter_dic = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0}
    chapter_text = ["", "", "", "", "", "", "", ""]

    prev_chapter_level = 0
    charter_no = ""

    doc = docx.Document(docx_filepath)

    for x, paragraph in enumerate(doc.paragraphs):
        if len(paragraph.text) <= 0:
            continue

        if paragraph.style != None and paragraph.style.name.find("스타일") >= 0 :
            nostyle = int(paragraph.style.name.replace("스타일",""))
            chapter_dic[nostyle] = chapter_dic[nostyle] + 1
            chapter_text[nostyle] = paragraph.text

            #현재레벨이 끝났다. 다시 레벨이 줄어든다.
            if prev_chapter_level == 3 and nostyle == 0:
                print("here")
            if prev_chapter_level > nostyle :
                for i in range(nostyle, 8):
                    chapter_dic[i+1] = 0

            charter_no = ""
            for no in range(0, nostyle+1):
                charter_no = charter_no + str(chapter_dic[no]) +"."
            #if charter_no < '6.':
            #    print('here')
            if chapter_dic[0] >=6 and prev_chapter_level >= 2 and nostyle <=2 :
                row_count = 1
                max_row_rount = 1
                for i, col in enumerate(col_contents):
                    if len(col) > 1:
                        print("{} : {}".format(i,col))
                        ws.cell(row=xrow, column=i + 1).value = col
                        ws.cell(row=xrow, column=i + 1).font = Font(size=10)
                        ws.cell(row=xrow, column=i + 1).alignment= Alignment(wrap_text=True, horizontal='left', vertical='center')

                        row_count = col.count('\n')
                        if max_row_rount < row_count :
                            max_row_rount = row_count
                ws.row_dimensions[xrow].height = (max_row_rount+1) * 15
                xrow = xrow + 1

            chapterSplit(charter_no, paragraph.text, prev_chapter_level)
            # 잘지워야 한다. 여기 중요함,
            if prev_chapter_level >= 2 and nostyle <= 2:
                for i, col in enumerate(col_contents):
                    #id는 chapterSplit에서 채웠다. 유효하다.
                    if len(col) > 1 and i > cols_idx['REQID']:
                        col_contents[i] = ""

            prev_chapter_level = nostyle
            #if charter_no =="3.1.4.3.":
            #    print("here")
            print(str(x) + ": "+charter_no+ ": "+ paragraph.text+"->"+ paragraph.style.name)
        #else :
        #    print(str(x) + "*****" + paragraph.text)




def chapterSplit( str_chap_no, strline, prev_lv ):
    chap_no = str_chap_no.split('.')
    chap_lv = len(chap_no) - 2
    global detail_req_col_idx
    global col_contents
    newline =False
    prefix_line = ""

    #if chap_no[chap_lv] == '1':
    #    print(f"newline {chap_lv} {strline}")
    if str_chap_no == '5.1.1.2.1.':
        print('here')
    if chap_lv <=2 :
        col_contents[chap_lv] = strline
        if chap_lv == 2 :
            col_contents[cols_idx['REQID']] = str_chap_no

    elif chap_lv == 3 :
        if strline.find("REQID") >= 0:
            detail_req_col_idx = cols_idx['REQID']
        elif strline.find("내용") >= 0:
            detail_req_col_idx = cols_idx['내용']
        elif strline.find("제약사항") >= 0:
            detail_req_col_idx = cols_idx['제약사항']
        elif strline.find("비고") >= 0:
            detail_req_col_idx = cols_idx['비고']
        else:
            detail_req_col_idx = -1

    elif chap_lv >= 4 : #
        if chap_lv == 4:
            prefix_line = '▶ '
        elif chap_lv > 4:
            for i in range(4, chap_lv ):
                prefix_line = prefix_line + '    '
            if chap_lv == 5:
                prefix_line = prefix_line + '- '
            elif chap_lv == 6:
                prefix_line = prefix_line + '* '
            elif chap_lv == 7:
                prefix_line = prefix_line + '> '

        strline = prefix_line + strline
        if detail_req_col_idx >= 0 and len(strline) > 0:
            if chap_no[chap_lv] == '1' and chap_lv == 4:
                col_contents[detail_req_col_idx] = strline
            else:
                col_contents[detail_req_col_idx] = col_contents[detail_req_col_idx] + "\r\n" + strline


    #print("len: {}, {} {}".format(chap_lv,detail_req_col_idx, strline ))

read_doc()
wb.save(filename=excel_filepath)

print("--end--")