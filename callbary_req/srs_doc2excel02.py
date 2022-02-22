import os.path
import sys

import docx

from docx import Document
from docx.document import Document as _Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph
from docx.opc.constants import RELATIONSHIP_TYPE as RT

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
import argparse
from nvlogger import Logger




class tableColumnInfo:
    def __init__(self, index, field_name, cell_width,  descCol ):
        self.index = index
        self.field_name = field_name
        self.contents = ""
        self.cell_width = cell_width
        self.descCol = descCol    #요구사항의 상세설명 여부

class chapterLevelInfo:
    def __init__(self):
        self.index = 0
        self.line = ""


class convertDoc2Excel:
    def __init__(self , table_col_list, table_col_map, chapLvList):
        self.table_col_list = table_col_list
        self.table_col_map = table_col_map
        self.chap_lv_list = chapLvList
        self.detail_req_col_idx = -1
        self.directory_path = 'D:/Documents/++ST++/00_PROOM2022/01.화물중계플랫폼/01.요구사항'
        self.docx_filename = '화물중계플랫폼2_요구사항_20220222_v04.docx'
        #self.docx_filepath = os.path.join(self.directory_path, self.docx_filename)
        self.docx_filepath = args.wordfilepath
        self.excel_filepath = self.docx_filepath.replace("docx", "xlsx")
        self.chapter_level_count = len(chapLvList)

        self.start_of_req_chapter_no = 5 #요구사항이 시작하는 lv1의 chapter no ********
        self.title_chapter_lv = 2 # 문서구조에서 style 0,1,2 는 고정됨. 이건 건들지 말것
        self.head_desc_chapter_lv = 3 # 상세요구사항의 항목(REQID, 내용, 제약사항 .... )
        self.detail_desc_chapter_lv = 4 # 상세요구사항 설명(

        self.reqid_map = {}
        self.reqid_duplicated = []
        self.reqid_col_index = self.getIndexColumnByName('REQID')

    def excelWorkbookInit(self):
        self.wb = openpyxl.Workbook()
        # 워크북을 생성하면 그 안에 워크시트 1개가 자동으로 생성
        self.ws = self.wb.active
        # 활성화 된 워크시트를 가리킴
        # wb.create_sheet('TITLE',0)

        #column이름 cell에 입력
        self.xrow = 1
        for i, col in enumerate(self.table_col_list) :
            self.ws.cell(row=self.xrow, column=col.index+1).value = col.field_name
            self.ws.column_dimensions[get_column_letter(col.index+1)].width = col.cell_width

        self.xrow = self.xrow + 1

    def convert(self):
        prev_chapter_level = 0
        doc = docx.Document(self.docx_filepath)

        self.excelWorkbookInit()

        for x, paragraph in enumerate(doc.paragraphs):
            if len(paragraph.text) <= 0:
                continue

            if paragraph.style != None and paragraph.style.name.find("스타일") >= 0 :
                nostyle = int(paragraph.style.name.replace("스타일",""))
                self.chap_lv_list[nostyle].index = self.chap_lv_list[nostyle].index + 1
                self.chap_lv_list[nostyle].line = paragraph.text

                #현재레벨이 끝났다. 다시 레벨이 줄어든다.
                if prev_chapter_level > nostyle :
                    for i in range(nostyle, self.chapter_level_count-1):
                        self.chap_lv_list[i+1].index = 0

                charter_no = ""
                for no in range(0, nostyle+1):
                    charter_no = charter_no + str(self.chap_lv_list[no].index) +"."
                #if charter_no < '6.':
                #    print('here')
                if self.chap_lv_list[0].index >= self.start_of_req_chapter_no and prev_chapter_level >= self.title_chapter_lv and nostyle <= self.title_chapter_lv :
                    self.writeExcelRow()

                self.chapterSplit(charter_no, paragraph.text, prev_chapter_level)
                # 잘지워야 한다. 여기 중요함,
                if prev_chapter_level >= self.title_chapter_lv and nostyle <= self.title_chapter_lv:
                    for i, col in enumerate(self.table_col_list):
                        #id는 chapterSplit에서 채웠다. 유효하다.
                        if len(col.contents) > 1 and col.descCol:
                            col.contents = ""

                prev_chapter_level = nostyle
                #if charter_no =="3.1.4.3.":
                #    print("here")
                logger.info(str(x) + ": "+charter_no+ ": "+ paragraph.text+"->"+ paragraph.style.name)
            #else :
            #    print(str(x) + "*****" + paragraph.text)
        self.writeExcelRow()
        self.wb.save(filename=self.excel_filepath)

    def writeExcelRow(self):
        row_count = 1
        max_row_rount = 1
        for i, col in enumerate(self.table_col_list):
            if len(col.contents) > 1:
                logger.info("{} : {}".format(i, col.contents))
                self.ws.cell(row=self.xrow, column=i + 1).value = col.contents
                self.ws.cell(row=self.xrow, column=i + 1).font = Font(size=10)
                self.ws.cell(row=self.xrow, column=i + 1).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                                vertical='center')

                row_count = col.contents.count('\n')
                if max_row_rount < row_count:
                    max_row_rount = row_count
        self.ws.row_dimensions[self.xrow].height = (max_row_rount + 1) * 15
        self.xrow = self.xrow + 1

    def setRequestID(self, request_id, doc_id ):
        if request_id in self.reqid_map:
            self.reqid_duplicated.append(request_id +":"+ doc_id)
        else:
            self.reqid_map[request_id] = doc_id

    def getIndexColumnByName(self, name):
        for col in self.table_col_list:
            if col.field_name.find(name) >= 0:
                return col.index
        return -1

    def chapterSplit(self, str_chap_no, strline, prev_lv ):
        chap_no = str_chap_no.split('.')
        chap_lv = len(chap_no) - 2
        newline =False
        prefix_line = ""

        #if str_chap_no == '5.1.1.2.1.':
        #    print('here')
        if chap_lv <= self.title_chapter_lv :
            self.table_col_list[chap_lv].contents = strline
            if chap_lv == self.title_chapter_lv :
                self.table_col_map['DOCID'].contents = str_chap_no

        elif chap_lv == self.head_desc_chapter_lv :
            self.detail_req_col_idx = -1
            for col in self.table_col_list:
                if col.descCol == True:
                    if strline.find(col.field_name) >= 0:
                        self.detail_req_col_idx = self.table_col_map[col.field_name].index


        elif chap_lv >= self.detail_desc_chapter_lv : #
            if chap_lv == self.detail_desc_chapter_lv:
                prefix_line = '▶ '
            elif chap_lv > self.detail_desc_chapter_lv:
                for i in range(self.detail_desc_chapter_lv, chap_lv ):
                    prefix_line = prefix_line + '    '
                if chap_lv == self.detail_desc_chapter_lv+1:
                    prefix_line = prefix_line + '- '
                elif chap_lv == self.detail_desc_chapter_lv+2:
                    prefix_line = prefix_line + '* '
                elif chap_lv == self.detail_desc_chapter_lv+3:
                    prefix_line = prefix_line + '> '

            strline = prefix_line + strline
            if self.detail_req_col_idx >= 0 and len(strline) > 0:
                if chap_no[chap_lv] == '1' and chap_lv == self.detail_desc_chapter_lv:
                    self.table_col_list[self.detail_req_col_idx].contents = strline
                else:
                    self.table_col_list[self.detail_req_col_idx].contents = self.table_col_list[self.detail_req_col_idx].contents + "\n" + strline

            if self.detail_req_col_idx == self.reqid_col_index:
                self.setRequestID(strline.strip(), self.table_col_map['DOCID'].contents)


        #print("len: {}, {} {}".format(chap_lv,self.detail_req_col_idx, strline ))
    def report(self):
        if len(self.reqid_duplicated)==0:
            logger.info("--There is no duplicated req_id ")
        else:
            logger.info("--Duplicated req_id count={}".format(len(self.reqid_duplicated)))
            logger.info("-- DUPLICATED REQUIRMENT ID ")
            for i, reqid in enumerate(self.reqid_duplicated):
                logger.info ("{}) reqid:{}".format(i, reqid))



def run():

    col_info_list = []
    col_info_list.append(tableColumnInfo(index=0, field_name='CHAP1', cell_width=20,  descCol=False))
    col_info_list.append(tableColumnInfo(index=1, field_name='CHAP2', cell_width=20,  descCol=False))
    col_info_list.append(tableColumnInfo(index=2, field_name='TITLE', cell_width=30,  descCol=False))
    col_info_list.append(tableColumnInfo(index=3, field_name='DOCID', cell_width=10,  descCol=False))
    col_info_list.append(tableColumnInfo(index=4, field_name='REQID', cell_width=15,  descCol=True))
    col_info_list.append(tableColumnInfo(index=5, field_name='담당자', cell_width=15, descCol=True))
    col_info_list.append(tableColumnInfo(index=6, field_name='내용', cell_width=90,  descCol=True))
    col_info_list.append(tableColumnInfo(index=7, field_name='제약사항', cell_width=30,  descCol=True))
    col_info_list.append(tableColumnInfo(index=8, field_name='문의사항', cell_width=30,  descCol=True))
    col_info_list.append(tableColumnInfo(index=9, field_name='검증방법', cell_width=30,  descCol=True))
    col_info_list.append(tableColumnInfo(index=10, field_name='비고', cell_width=50,  descCol=True))

    col_map_by_fieldname ={}
    for _, col_info in enumerate(col_info_list):
        col_map_by_fieldname[col_info.field_name] = col_info

    chapLvList = []
    max_chapter_lv_cnt = 8
    for chapLv in range(0, max_chapter_lv_cnt+1):
        chapLvList.append(chapterLevelInfo())

    converter = convertDoc2Excel(col_info_list,col_map_by_fieldname,chapLvList)
    converter.convert()
    converter.report()


if __name__ == '__main__':
    global args
    global logger

    try:
        logManager = Logger.instance()
        logManager.setLogger("cb.log")
        logger = logManager.getLogger()

        argParser = argparse.ArgumentParser()
        argParser.add_argument("--wordfilepath", help="word file path", required=True, default='srs.docx',
                            type=str)

        args = argParser.parse_args()
        if os.path.isfile(args.wordfilepath) == False:
            logger.info("invalid file path:{}".format(args.wordfilepath))
            sys.exit("invalid word document file:"+args.wordfilepath)
        else:
            logger.info("file path:{}".format(args.wordfilepath))

        logger.info("--start--")
        run()
        logger.info("--end--")
    except Exception as err:
        logger.info("error:"+err)
